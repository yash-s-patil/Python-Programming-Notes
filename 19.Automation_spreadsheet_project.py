"""
project - Automation with python ( Working with spreadsheets)
Read Spreadsheet file and automate stuff
Task:
1. List each company with respective product count
2. List product with inventory less than 10
3. List each company with respective total inventory value
4. Calculate and write inventory value for each product to spreadsheet.
"""
# pip install openpyxl ( to work with excel)
import openpyxl
# load workbook
inv_file = openpyxl.load_workbook("17.inventory.xlsx")
product_list = inv_file["Sheet1"]


products_per_supplier = {}
total_value_per_supplier = {}
products_under_10_inv ={}

for product_row in range(2, product_list.max_row + 1):
    supplier_name = product_list.cell(product_row, 4).value
    inventory = product_list.cell(product_row, 2).value
    price = product_list.cell(product_row, 3).value
    product_num = product_list.cell(product_row, 1).value
    inventory_price = product_list.cell(product_row, 5)

    # Task 1 : List each company with respective product count
    if supplier_name in products_per_supplier:
        products_per_supplier[supplier_name] = products_per_supplier[supplier_name] + 1
    else:
        products_per_supplier[supplier_name] = 1

    # Task 2: List each company with respective total inventory value
    if supplier_name in total_value_per_supplier:
        total_value_per_supplier[supplier_name] = total_value_per_supplier[supplier_name] + inventory * price
    else:
        total_value_per_supplier[supplier_name] = inventory * price

    # Task 3: List product with inventory less than 10
    if inventory < 10:
        products_under_10_inv[int(product_num)] = int(inventory)

    # Task 4: Calculate and write inventory value for each product to spreadsheet and save it
    inventory_price.value = inventory * price


print(products_per_supplier)
print(total_value_per_supplier)
print(products_under_10_inv)
inv_file.save("18.new_inventory.xlsx")

